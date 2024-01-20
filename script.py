import os
import sys
import time
import random

# check if all required packages are installed
try:
    import requests
    import pandas as pd
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Color, Alignment
except ImportError:
    print(f"{"\033[91m"}ERROR: The following packages are required: {"\033[1m"}requests, pandas, bs4, openpyxl{"\033[0m"}")
    print("Please install with: pip install -r requirements.txt")
    exit()

# [REQUIRED] Replace with actual urls. HINT: only domains names have to be changed
##############################################
list_url        = "https://creamwhip.com/games-list/"
entry_url       = "https://creamwhip.com"
critic_site_url = "https://www.criticsite.com/game/"
##############################################

# [OPTIONAL] Preferences
# -------------------------------------------

# Specify the name of the local gamelist    | default: "gamelist.html"
file_name             = "gamelist.html"

# Skip rescraping scores (True/False)       | default: True
# Note: this saves the scores into a csv file, delete the file to rescrape
skip_rescraping       = True

# Output with timestamp  (True/False)       | default: True
save_with_timestamp   = True             

# Specify wait time inbetween HTTP requests | default: 2-10
# Note: requests are randomized ([min, max] in sec.)
wait_between_requests = [2, 10]

# Add cool link symbol in spreadsheet       | default: True
enable_link_symbol    = False

# Here you can specify the CSS selectors for the scores and release year
# Note: you can find the selectors by inspecting the HTML of the page and 
#       right-clicking on the element, then selecting "Copy" -> "Copy selector"
class Selectors:
    # criticscore should look like this at the end: ...<span>90</span>
    criticscore = "#__layout > div > div.c-layoutDefault_page > div.c-pageProductGame > div:nth-child(1) > div > div > div.c-productHero_player-scoreInfo.u-grid.g-grid-container > div.c-productHero_score-container.u-flexbox.u-flexbox-column.g-bg-white > div.c-productHero_scoreInfo.g-inner-spacing-top-medium.g-outer-spacing-bottom-medium.g-outer-spacing-top-medium > div:nth-child(1) > div > div.c-productScoreInfo_scoreContent.u-flexbox.u-flexbox-alignCenter.u-flexbox-justifyFlexEnd.g-width-100.u-flexbox-nowrap > div.c-productScoreInfo_scoreNumber.u-float-right > div > div > span"

    # example: ...<span>8.5</span>
    userscore = "#__layout > div > div.c-layoutDefault_page > div.c-pageProductGame > div:nth-child(1) > div > div > div.c-productHero_player-scoreInfo.u-grid.g-grid-container > div.c-productHero_score-container.u-flexbox.u-flexbox-column.g-bg-white > div.c-productHero_scoreInfo.g-inner-spacing-top-medium.g-outer-spacing-bottom-medium.g-outer-spacing-top-medium > div.c-productScoreInfo.u-clearfix > div.c-productScoreInfo_scoreContent.u-flexbox.u-flexbox-alignCenter.u-flexbox-justifyFlexEnd.g-width-100.u-flexbox-nowrap > div.c-productScoreInfo_scoreNumber.u-float-right > div > div > span"

    # example: ...<span>2019</span>
    r_year = "#__layout > div > div.c-layoutDefault_page > div.c-pageProductGame > div:nth-child(1) > div > div > div.c-productHero_player-scoreInfo.u-grid.g-grid-container > div.c-productHero_score-container.u-flexbox.u-flexbox-column.g-bg-white > div.g-text-xsmall > span.u-text-uppercase"

# -------------------------------------------
    

class ConsoleHelpers:
    # C_ -> Colors
    # from: https://stackoverflow.com/questions/287871/how-do-i-print-colored-text-to-the-terminal
    C_HEADER = "\033[95m"
    C_OKBLUE = "\033[94m"
    C_OKCYAN = "\033[96m"
    C_OKGREEN = "\033[92m"
    C_WARNING = "\033[93m"
    C_FAIL = "\033[91m"
    C_END = "\033[0m"

    # F_ -> Formatting
    F_BOLD = "\033[1m"
    F_UNDERLINE = "\033[4m"
    F_BLANK_LINE = "\033[K"

if list_url == "https://creamwhip.com/games-list/" or entry_url == "https://creamwhip.com":
    print("Please specify the correct URLs in the script.py (line 22, 23)")
    exit()

# Check if the local file exists
if os.path.exists(file_name):
    # Open the local file and read its content
    with open(file_name, "r", encoding="utf-8") as f:
        html_content = f.read()
else:
    # Send HTTP request to the specified URL and save the response from server
    r = requests.get(list_url)
    html_content = r.text

    # Save the text of the webpage to a local file
    with open(file_name, "w", encoding="utf-8") as f:
        f.write(html_content)


soup = BeautifulSoup(html_content, "html.parser")

# Check if the CSV file exists and read it
if os.path.exists("scores.csv") and skip_rescraping:
    scores_df = pd.read_csv("scores.csv")
else:
    scores_df = pd.DataFrame(columns=["Name", "Link", "Critic Score", "User Score", "Release Year"])

print("Looking for games...")

for item in soup.find_all("li", {"class": "az-list-item"}):
    link_tag = item.find("a")
    name = link_tag.text.split(" Free Download")[0]
    link = entry_url + link_tag.get("href")

    # Check if the scores for the game are in the CSV file
    scores = scores_df.loc[scores_df["Name"] == name, ["Critic Score", "User Score", "Release Year"]]
    if scores.size == 0:
        print("=============================")
        # Wait for specified time, so that the server doesn"t block the requests
        wait_time = random.randint(wait_between_requests[0], wait_between_requests[1])
        for i in range(wait_time):
            print(f"  Waiting (" + str(wait_time - i) + "s) for the server to respond...", end="\r")
            time.sleep(1)

        sys.stdout.write(ConsoleHelpers.F_BLANK_LINE)
        print(f"  Requesting...", end="\r")

        # Scrape the critic site score, user score, and release year
        criticsite_url = critic_site_url+f"{name.replace(":", "").replace("!", "").replace("’", "").replace("'", "").replace(".", "").replace("(", "").replace(")", "").replace(" – ", "").replace(" GOTY", "").replace(" ", "-").replace("---", "").replace("‘", "").lower()}"
        
        headers = {"User-Agent": "MyBot/1.0"}
        response = requests.get(criticsite_url, headers=headers)

        if (response.status_code == 403):
            print(name, ":", "N/A | N/A - N/A")
            print(f"{ConsoleHelpers.C_FAIL}FAIL: {response.status_code} - {criticsite_url}")
            print(f"Server doesn't authorize traffic! Please wait a few minutes and try again.{ConsoleHelpers.C_END}")
            exit()

        if (response.status_code != 200):
            print(name, ":", "N/A | N/A - N/A")
            print(f"{ConsoleHelpers.C_WARNING}ERROR: {response.status_code} - {criticsite_url}{ConsoleHelpers.C_END}")
            continue

        criticsite_soup = BeautifulSoup(response.text, "html.parser")
        criticscore    = criticsite_soup.select_one(Selectors.criticscore)
        userscore    = criticsite_soup.select_one(Selectors.userscore)
        release_year = criticsite_soup.select_one(Selectors.r_year)
        
        criticscore    = criticscore.text if criticscore is not None else "N/A"
        userscore    = userscore.text if userscore is not None else "N/A"
        release_year = release_year.text.split(", ")[1] if release_year is not None else "N/A"

        sys.stdout.write(ConsoleHelpers.F_BLANK_LINE)
        if criticscore != "N/A" and userscore != "N/A" and release_year != "N/A":
            print(name, ":", criticscore, "|", userscore, "-", release_year, "✓")
            print(f"{ConsoleHelpers.C_OKGREEN}GREAT SUCCESS: Everything was found on page!{ConsoleHelpers.C_END}")
        else :
            print(name, ":", criticscore, "|", userscore, "-", release_year)
            print(f"{ConsoleHelpers.C_OKCYAN}SUCCESS: At least the page was found.{ConsoleHelpers.C_END}")

        # Add the scores and release year to the DataFrame
        scores_df.loc[len(scores_df)] = [name, link, criticscore, userscore, release_year]

        # Save the DataFrame to the CSV file
        scores_df.to_csv("scores.csv", index=False)

print("=============================")
print("Cooking up Excel file...")

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create a dictionary to store the names and links
name_link_dict = pd.Series(scores_df.Link.values,index=scores_df.Name).to_dict()

# Remove the "Link" column from the DataFrame
scores_df = scores_df.drop(columns=["Link"])

# Convert the scores to numeric values
scores_df["Critic Score"] = pd.to_numeric(scores_df["Critic Score"], errors="coerce")
scores_df["Critic Score"] = scores_df["Critic Score"].fillna("")

scores_df["User Score"] = pd.to_numeric(scores_df["User Score"], errors="coerce")
scores_df["User Score"] = scores_df["User Score"].fillna("")

# Append the rows of the DataFrame to the worksheet
for r in dataframe_to_rows(scores_df, index=False, header=True):
    ws.append(r)

# Set the height of all rows and center the text vertically
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18
    for cell in row:
        cell.alignment = Alignment(vertical="center")

# Auto-size the columns
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Change the format of the name cells to be recognized as links by Excel
blue_font = Color("0563C1")
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        link = name_link_dict[cell.value]  # Use the dictionary to get the link
        cell.hyperlink = link
        cell.font = Font(underline="single", color=blue_font)  # Add underline
        if enable_link_symbol: cell.value = cell.value + " 🔗"  # Add link symbol at the end

# save to excel file with timestamp
if (save_with_timestamp):
    import datetime
    now = datetime.datetime.now()
    timestamp = now.strftime("%y%m%d_%H%M")
    wb.save(f"output_{timestamp}.xlsx")
else:
    wb.save("output.xlsx")

print("Done!")
